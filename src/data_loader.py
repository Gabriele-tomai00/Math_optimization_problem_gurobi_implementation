import openpyxl
import os


def find_label(ws, label):
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            val = ws.cell(row=r, column=c).value
            if val and label in str(val).upper():
                return r, c
    raise ValueError(f"{label} not found")


def read_block(ws, start_row, start_col, is_vector=False):
    data = []
    row = start_row + 1
    col_start = start_col + 1

    while True:
        col = col_start
        row_data = []

        cell_value = ws.cell(row=row, column=col).value

        # Stop if first cell is empty
        if cell_value is None:
            break

        while True:
            cell_value = ws.cell(row=row, column=col).value

            if isinstance(cell_value, int) and cell_value >= 0:
                row_data.append(cell_value)
                col += 1
            else:
                break

        if is_vector:
            data.append(row_data[0])
        else:
            data.append(row_data)

        row += 1

    return data

def read_penalty(ws):
    found_capacity = False
    found_empty = False

    for row_idx in range(1, ws.max_row + 1):
        cell_value = ws.cell(row=row_idx, column=1).value

        # Find CAPACITY
        if cell_value and "CAPACITY" in str(cell_value).upper():
            found_capacity = True
            continue

        # After CAPACITY
        if found_capacity:
            if cell_value is None:
                found_empty = True
                continue

            if found_empty and isinstance(cell_value, int) and cell_value % 100 == 0:
                return cell_value

    raise ValueError("Penalty not found")

def read_from_excel(ws, label, is_vector=False):

    r, c = find_label(ws, label)
    return read_block(ws, r, c, is_vector)

def pretty_print(name, data):
    print(f"{name}:")
    
    if isinstance(data[0], list):  # matrix
        for row in data:
            print("  ", row)
    else:  # vector
        for i, val in enumerate(data, 1):
            print("  ", val)    


def get_data_from_file_excel(file_path, sheet_index=None):
    wb = openpyxl.load_workbook(file_path, data_only=True)

    if sheet_index is not None:
        ws = wb.worksheets[sheet_index]
        sheet_name = ws.title

        if not sheet_name.isdigit():
            return {}

        demand = read_from_excel(ws, "DAMEND")
        capacity = read_from_excel(ws, "CAPACITY")
        cost = read_from_excel(ws, "COST")
        price = read_from_excel(ws, "PRICE")
        revenue = read_from_excel(ws, "REVENUE", is_vector=True)
        penalty = read_penalty(ws)

        return {
            "demand": demand,
            "capacity": capacity,
            "cost": cost,
            "price": price,
            "revenue": revenue,
            "penalty": penalty
        }

    # fallback: tutti i fogli
    all_instances = {}
    for ws in wb.worksheets:
        sheet_name = ws.title
        if not sheet_name.isdigit():
            continue

        demand = read_from_excel(ws, "DAMEND")
        capacity = read_from_excel(ws, "CAPACITY")
        cost = read_from_excel(ws, "COST")
        price = read_from_excel(ws, "PRICE")
        revenue = read_from_excel(ws, "REVENUE", is_vector=True)
        penalty = read_penalty(ws)

        all_instances[sheet_name] = {
            "demand": demand,
            "capacity": capacity,
            "cost": cost,
            "price": price,
            "revenue": revenue,
            "penalty": penalty
        }

    return all_instances

def validate_dimensions(Q, C, c, p, R):
    # print(f"DEBUG Dimensioni:")
    # print(f"  Hotel (I): {len(R)}")
    # print(f"  Nodi Domanda (J): {len(Q)}")
    # print(f"  Tipi Stanza (K): {len(Q[0]) if Q else 0}")
    # print(f"  Matrice Costi (c): {len(c)} righe, {len(c[0]) if c else 0} colonne")
    # print(f"  Matrice Prezzi (p): {len(p)} righe, {len(p[0]) if p else 0} colonne")
    
    # Number of hotels (I)
    I = len(R)

    errors = []

    if len(c) != I:
        errors.append(f"COST rows ({len(c)}) != number of hotels ({I})")

    if len(p) != I:
        errors.append(f"PRICE rows ({len(p)}) != number of hotels ({I})")

    if len(C) != I:
        errors.append(f"CAPACITY rows ({len(C)}) != number of hotels ({I})")

    # Optional extra checks (very useful)
    if len(Q) == 0:
        errors.append("DEMAND matrix is empty")
        return False
    else:
        K = len(Q[0])
        for i, row in enumerate(c):
            if len(row) != K:
                errors.append(f"COST row {i} has {len(row)} columns, expected {K}")
        for i, row in enumerate(p):
            if len(row) != K:
                errors.append(f"PRICE row {i} has {len(row)} columns, expected {K}")

    if errors:
        return False

    print("Validation OK: dimensions are consistent.")
    return True


def get_data_for_model(file_name, sheet_index=None):
    file_path = os.path.join("..", "quarantine_hotel_instances", file_name)

    if not os.path.exists(file_path):
        print(f"Errore: File {file_path} non trovato.")
        return

    print(f"Caricamento dati da {file_name}...")
    data = get_data_from_file_excel(file_path, 0)

    if not validate_dimensions(
                data["demand"],
                data["capacity"],
                data["cost"],
                data["price"],
                data["revenue"] 
            ):
        print("Error: Dimension validation failed. Please check the input data.")
        exit(1)

    # --- CLEAN AND VALIDATE DATA ---
    # Rimuoviamo righe vuote e assicuriamoci che siano liste di liste
    Q = [row for row in data["demand"] if len(row) > 0]
    C = [row for row in data["capacity"] if len(row) > 0]
    c = [row for row in data["cost"] if len(row) > 0]
    p = [row for row in data["price"] if len(row) > 0]
    R = [val for val in data["revenue"] if val is not None]
    gamma = data["penalty"]
    return Q, C, c, p, R, gamma


if __name__ == "__main__":

    all_files_data = {}

    for i in range(1, 17):
        file_name = f"{i}.xlsx"
        file_path = os.path.join("..", "quarantine_hotel_instances", file_name)

        if not os.path.exists(file_path):
            print(f"File {file_name} non trovato, salto")
            continue

        print(f"\n=== Processing {file_name} ===")

        instances = get_data_from_file_excel(file_path)
        all_files_data[file_name] = instances

    file_name = f"1.xlsx"
    file_path = os.path.join("..", "quarantine_hotel_instances", file_name)

    if not os.path.exists(file_path):
        print(f"File {file_name} not found")
        exit(1)

    print(f"\n=== Processing {file_name} ===")

    data = get_data_from_file_excel(file_path, 0)

    # --- PULIZIA E VALIDAZIONE DATI ---
    # Rimuoviamo righe vuote e assicuriamoci che siano liste di liste
    Q = [row for row in data["demand"] if len(row) > 0]
    C = [row for row in data["capacity"] if len(row) > 0]
    c = [row for row in data["cost"] if len(row) > 0]
    p = [row for row in data["price"] if len(row) > 0]
    R = [val for val in data["revenue"] if val is not None]
    gamma = data["penalty"]
    
    if not validate_dimensions(
                data["demand"],
                data["capacity"],
                data["cost"],
                data["price"],
                data["revenue"] 
            ):
        print("Error: Dimension validation failed. Please check the input data.")
        exit(1)

    pretty_print("DEMAND", Q)
    pretty_print("CAPACITY", C)
    pretty_print("COST", c)
    pretty_print("PRICE", p)
    pretty_print("REVENUE", R)
    pretty_print("PENALTY", [gamma])

    # Debug delle dimensioni (Fondamentale per l'IndexError)
    print(f"DEBUG Dimensioni:")
    print(f"  Hotel (I): {len(R)}")
    print(f"  Nodi Domanda (J): {len(Q)}")
    print(f"  Tipi Stanza (K): {len(Q[0]) if Q else 0}")
    print(f"  Matrice Costi (c): {len(c)} righe, {len(c[0]) if c else 0} colonne")
    print(f"  Matrice Prezzi (p): {len(p)} righe, {len(p[0]) if p else 0} colonne")

